import os
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.core.config import settings

def generate_escala_geral_excel(titulo: str, mes_ano: str, payload: dict, filename: str) -> str:
    """
    Generates Excel spreadsheet matching Image 1 layout (Purple theme, merged headers, DATA & DIA rows).
    """
    output_dir = os.path.join(settings.UPLOAD_DIR, "exportacoes")
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Escala Geral"

    cultos = payload.get("cultos", [])
    funcoes = payload.get("funcoes", [])
    matriz = payload.get("matriz", {})

    purple_fill = PatternFill(start_color="C084FC", end_color="C084FC", fill_type="solid")
    purple_sub_fill = PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid")
    role_fill = PatternFill(start_color="E9D5FF", end_color="E9D5FF", fill_type="solid")

    font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
    font_cell = Font(name="Calibri", size=9, color="000000")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin', color='475569'),
        right=Side(style='thin', color='475569'),
        top=Side(style='thin', color='475569'),
        bottom=Side(style='thin', color='475569')
    )

    # Row 1: Banner Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cultos) + 1)
    cell_title = ws.cell(row=1, column=1, value=f"ESCALA DO MÊS DE {mes_ano.upper()}")
    cell_title.fill = purple_fill
    cell_title.font = Font(name="Calibri", size=14, bold=True, color="000000")
    cell_title.alignment = align_center

    # Row 2: Cult Types
    ws.cell(row=2, column=1, value="").fill = purple_sub_fill
    for idx, c in enumerate(cultos, 2):
        cell = ws.cell(row=2, column=idx, value=c.get("culto", ""))
        cell.fill = purple_sub_fill
        cell.font = font_bold
        cell.alignment = align_center

    # Row 3: DATA
    ws.cell(row=3, column=1, value="DATA").fill = purple_sub_fill
    ws.cell(row=3, column=1).font = font_bold
    ws.cell(row=3, column=1).alignment = align_center
    for idx, c in enumerate(cultos, 2):
        cell = ws.cell(row=3, column=idx, value=c.get("data", ""))
        cell.fill = purple_sub_fill
        cell.font = font_bold
        cell.alignment = align_center

    # Row 4: DIA
    ws.cell(row=4, column=1, value="DIA").fill = purple_sub_fill
    ws.cell(row=4, column=1).font = font_bold
    ws.cell(row=4, column=1).alignment = align_center
    for idx, c in enumerate(cultos, 2):
        cell = ws.cell(row=4, column=idx, value=c.get("dia", ""))
        cell.fill = purple_sub_fill
        cell.font = font_bold
        cell.alignment = align_center

    # Role Rows
    current_row = 5
    for funcao in funcoes:
        cell_role = ws.cell(row=current_row, column=1, value=funcao)
        cell_role.fill = role_fill
        cell_role.font = font_bold
        cell_role.alignment = align_center

        for cIdx, c in enumerate(cultos, 2):
            val = matriz.get(f"{cIdx-2}_{funcao}", "//")
            cell_val = ws.cell(row=current_row, column=cIdx, value=val or "//")
            cell_val.font = font_cell
            cell_val.alignment = align_center
        current_row += 1

    # Footer Obs Row
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(cultos) + 1)
    obs_cell = ws.cell(row=current_row, column=1, value="Obs.: TODOS OS CULTOS DEVEM SER PRECEDIDOS DE PELO MENOS 15 MINUTOS DE ORAÇÃO!")
    obs_cell.fill = purple_fill
    obs_cell.font = Font(name="Calibri", size=10, bold=True, color="000000")
    obs_cell.alignment = align_center

    # Add borders to all cells
    for row in ws.iter_rows(min_row=1, max_row=current_row, min_col=1, max_col=len(cultos)+1):
        for cell in row:
            cell.border = thin_border

    wb.save(filepath)
    return f"/uploads/exportacoes/{filename}"


def generate_escala_ebi_excel(titulo: str, mes_ano: str, payload: dict, filename: str) -> str:
    """
    Generates Excel spreadsheet matching Image 2 layout (Navy theme, Sunday columns, Class rows).
    """
    output_dir = os.path.join(settings.UPLOAD_DIR, "exportacoes")
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Escala EBI"

    domingos = payload.get("domingos", [])
    salas = payload.get("salas", [])
    matriz = payload.get("matriz", {})

    navy_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    sub_navy_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    label_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    font_white_bold = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
    font_cell = Font(name="Calibri", size=10, color="000000")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin', color='94A3B8'),
        right=Side(style='thin', color='94A3B8'),
        top=Side(style='thin', color='94A3B8'),
        bottom=Side(style='thin', color='94A3B8')
    )

    # Title Row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(domingos) + 1)
    cell_title = ws.cell(row=1, column=1, value="ESCOLA BÍBLICA INFANTIL")
    cell_title.fill = navy_fill
    cell_title.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    cell_title.alignment = align_center

    # Month Banner
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(domingos) + 1)
    cell_month = ws.cell(row=2, column=1, value=mes_ano.upper())
    cell_month.fill = sub_navy_fill
    cell_month.font = font_white_bold
    cell_month.alignment = align_center

    # Header Columns
    cell_corner = ws.cell(row=3, column=1, value="Salas / Datas")
    cell_corner.fill = sub_navy_fill
    cell_corner.font = font_white_bold
    cell_corner.alignment = align_center

    for idx, dom in enumerate(domingos, 2):
        cell = ws.cell(row=3, column=idx, value=dom)
        cell.fill = sub_navy_fill
        cell.font = font_white_bold
        cell.alignment = align_center

    # Data Rows
    current_row = 4
    for sala in salas:
        cell_label = ws.cell(row=current_row, column=1, value=sala.get("label", ""))
        cell_label.fill = label_fill
        cell_label.font = font_bold
        cell_label.alignment = align_center

        for dIdx, dom in enumerate(domingos, 2):
            val = matriz.get(f"{sala.get('key')}_{dIdx-2}", "-")
            cell_val = ws.cell(row=current_row, column=dIdx, value=val or "-")
            cell_val.font = font_cell
            cell_val.alignment = align_center
        current_row += 1

    # Format borders & column widths
    for row in ws.iter_rows(min_row=1, max_row=current_row-1, min_col=1, max_col=len(domingos)+1):
        for cell in row:
            cell.border = thin_border

    ws.column_dimensions['A'].width = 30
    for col_idx in range(2, len(domingos) + 2):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 22

    wb.save(filepath)
    return f"/uploads/exportacoes/{filename}"


def generate_excel_report(title: str, headers: list, data: list, filename: str) -> str:
    """
    Generates a styled Excel spreadsheet (.xlsx).
    """
    output_dir = os.path.join(settings.UPLOAD_DIR, "exportacoes")
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório"

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")

    ws.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    for row in data:
        ws.append(row)

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(filepath)
    return f"/uploads/exportacoes/{filename}"


def generate_csv_report(headers: list, data: list, filename: str) -> str:
    """
    Generates a CSV file.
    """
    output_dir = os.path.join(settings.UPLOAD_DIR, "exportacoes")
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)

    with open(filepath, mode="w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file, delimiter=";")
        writer.writerow(headers)
        writer.writerows(data)

    return f"/uploads/exportacoes/{filename}"
